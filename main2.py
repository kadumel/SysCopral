


import pandas as pd
import openpyxl as xl
from xlrd import open_workbook, xldate
from datetime import datetime
import copral.app.connectionFactory as cf
import os
import fnmatch

path = 'D:\Projetos\Copral\Importacao\LogPos_HXA9626 - ABR 2020.xls'


def getFileTela(cam):
    ListPath = []
    # path = 'C:\Ezbi ETL\SIMVAL\Arquivos'
    path = cam

    for files in os.walk(path):
        # print(files[2])
        for filename in fnmatch.filter(files[2], '*.xls'):
            ListPath.append(filename)
            print(filename)


    for i in range(len(ListPath)):
        importDados(path + '\\' + str(ListPath[i]))
        os.remove(path + '\\' + str(ListPath[i]))



def importDados(path):

    wb = open_workbook(path)
    ws = wb.sheet_by_index(0)

    listColCab = []
    countColCab = 0

    for i in ws.row(0):
        if i.value != '':
            listColCab.append(countColCab)
        countColCab += 1


    listColCorpo = []
    countColCorpo = 0

    for i in ws.row(2):
        if i.value != '' and i.value != 'Ponto de Referência':
            listColCorpo.append(countColCorpo)
        countColCorpo += 1



    listCab = []
    for i in listColCab:
        listCab.append(ws.cell(1,i).value)



    listCorpo = []
    n = ws.nrows
    for r in range(ws.nrows -3):
        dados = []
        for j in listColCorpo:
            print(j)
            if j == 1:
                a1_as_datetime = xldate.xldate_as_datetime(ws.cell(r + 3, j).value, wb.datemode)
                dados.append(a1_as_datetime)
            else:
                dados.append(ws.cell(r + 3, j).value)


        print(listCab)
        print(dados)
        listCorpo.append(listCab + dados + [datetime.today(), path, None])



    sql = 'insert into ImportadosExcel values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'



    cf.insertLote(sql, listCorpo)
    print(path)